
from django.shortcuts import render, redirect, get_object_or_404

from accounts.decorators import owner_or_staff_required

from accounts.models import User
from orders.models import Order, Coupon, OrderTimeline, ReturnRequest,Refund, Replacement, Exchange, ReturnComment
from django.contrib import messages

from store.models import Category, Product, Brand, ProductImage, AbandonedCart, Wishlist
from .forms import CategoryForm, ProductForm, BrandForm, ProductGalleryForm, CouponForm, AnnouncementForm, EmailCampaignForm

from .decorators import owner_required
from django.core.paginator import Paginator

from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.db.models.functions import TruncMonth

import json

from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from orders.models import ExchangeRate

from .models import FeatureToggle, Announcement, WebsiteSettings, CustomerActivity, EmailCampaign, EmailCampaignLog
from .helpers import clear_feature_cache
from django.views.decorators.http import require_POST
from django.db import transaction
from django.core.mail import EmailMultiAlternatives
from .email_utils import send_campaign_email
from django.core.mail import send_mail
from django.conf import settings
# Create your views here.


@owner_or_staff_required
def dashboard_home(request):

    total_orders = Order.objects.count()

    total_categories = Category.objects.count()

    total_customers = User.objects.filter(
        role="customer",
    ).count()

    total_products = Product.objects.count()

    revenue = (
        Order.objects.filter(
            status="delivered",
            paid=True,
        ).aggregate(
            total=Sum("final_total")
        )["total"] or 0
    )

    today = timezone.now().date()

    today_orders = Order.objects.filter(
        created_at__date=today,
    ).count()

    today_revenue = (
        Order.objects.filter(
            created_at__date=today,
            paid=True,
            status="delivered",
        ).aggregate(
            total=Sum("final_total"),
        )["total"] or 0
    )

    monthly_revenue = (

        Order.objects.filter(

            paid=True,

            status="delivered",

        )

        .annotate(

            month=TruncMonth("created_at"),

        )

        .values(

            "month",

        )

        .annotate(

            revenue=Sum("final_total"),

        )

        .order_by(

            "month",

        )

    )

    monthly_orders = (

        Order.objects.annotate(

            month=TruncMonth(

                "created_at"

            )

        )

        .values(

            "month"

        )

        .annotate(

            total=Count("id")

        )

        .order_by(

            "month"

        )

    )

    top_products = (
        Product.objects.only(
            "id",
            "name",
        )
        .annotate(
            total_sold=Sum("orderitem__quantity")
        )
        .order_by("-total_sold")[:5]
    )

    top_categories = (
        Category.objects.only(
            "id",
            "name",
        )
        .annotate(
            total_sold=Sum("products__orderitem__quantity")
        )
        .order_by("-total_sold")[:5]
    )


    recent_orders = Order.objects.select_related(
        "user",
    ).order_by(
        "-created_at",
    )[:5]

    recent_customers = User.objects.filter(
        role="customer",
    ).only(
        "id",
        "username",
        "email",
        "date_joined",
    ).order_by("-date_joined")[:5]

    low_stock_products = Product.objects.filter(
        stock__gt=0,
        stock__lte=5,
    ).only(
        "id",
        "name",
        "stock",
    )

    out_of_stock_products = Product.objects.filter(
        stock=0,
    ).only(
        "id",
        "name",
        "stock",
    )


    monthly_labels = [
        item["month"].strftime("%b %Y")
        for item in monthly_revenue
    ]

    monthly_revenue_data = [
        float(item["revenue"] or 0)
        for item in monthly_revenue
    ]

    monthly_order_data = [
        item["total"]
        for item in monthly_orders
    ]

    top_product_labels = [
        product.name
        for product in top_products
    ]

    top_product_data = [
        product.total_sold or 0
        for product in top_products
    ] 

    top_category_labels = [
        category.name
        for category in top_categories
    ]

    top_category_data = [
        category.total_sold or 0
        for category in top_categories
    ]


    payment_methods = (
        Order.objects.filter(
            paid=True,
            status="delivered",
        )
        .values("payment_method")
        .annotate(
            total=Sum("final_total"),
        )
    )

    payment_labels = [
        item["payment_method"].upper()
        for item in payment_methods
    ]

    payment_data = [
        float(item["total"] or 0)
        for item in payment_methods
    ]

    payment_method_labels = [
        item["payment_method"].upper()
        for item in payment_methods
    ]

    payment_method_data = [
        float(item["total"] or 0)
        for item in payment_methods
    ]

    

    best_customers = (

        User.objects.filter(
            role="customer",
        )

        .annotate(

            total_spent=Sum(
                "orders__final_total",
                filter=Q(
                    orders__paid=True,
                    orders__status="delivered",
                ),
            ),

            total_orders=Count(
                "orders",
                filter=Q(
                    orders__paid=True,
                    orders__status="delivered",
                ),
            ),

        )

        .filter(
            total_spent__isnull=False,
        )

        .order_by(
            "-total_spent",
        )[:5]

    )

    context = {

        "total_orders": total_orders,

        "total_categories": total_categories,

        "total_customers": total_customers,

        "total_products": total_products,

        "revenue": revenue,

        "today_orders": today_orders,

        "today_revenue": today_revenue,

        "monthly_revenue": monthly_revenue,

        "monthly_orders": monthly_orders,

        "top_products": top_products,

        "top_categories": top_categories,

        "recent_orders": recent_orders,

        "recent_customers": recent_customers,

        "low_stock_products": low_stock_products,

        "out_of_stock_products": out_of_stock_products,

        "low_stock_count": low_stock_products.count(),

        "out_of_stock_count": out_of_stock_products.count(),

        "monthly_labels": json.dumps(monthly_labels),
        "monthly_revenue_data": json.dumps(monthly_revenue_data), 
        "monthly_order_data": json.dumps(monthly_order_data),

        "top_product_labels": json.dumps(top_product_labels),
        "top_product_data": json.dumps(top_product_data),

        "top_category_labels": json.dumps(top_category_labels),
        "top_category_data": json.dumps(top_category_data),

        "payment_labels": json.dumps(payment_labels),

        "payment_data": json.dumps(payment_data),

        "payment_method_labels": json.dumps(payment_method_labels),
        "payment_method_data": json.dumps(payment_method_data),

        "best_customers": best_customers,

    }

    if request.user.role == "owner":

        return render(
            request,
            "dashboard/owner_dashboard.html",
            context,
        )

    if request.user.role == "staff":

        return render(
            request,
            "dashboard/staff_dashboard.html",
            context,
        )

    return redirect("home")


@owner_or_staff_required
def reports(request):

    revenue_by_payment = (

        Order.objects.filter(

            paid=True,
            status="delivered",

        )

        .values("payment_method")

        .annotate(

            total=Sum("final_total"),

            orders=Count("id"),

        )

    )

    best_customers = (

        User.objects.filter(

            role="customer",

        )

        .annotate(

            total_spent=Sum("orders__final_total"),

            total_orders=Count("orders"),

        )

        .order_by(

            "-total_spent",

        )[:10]

    )

    low_stock = Product.objects.filter(
        stock__lte=5,
    ).select_related(
        "category",
    )

    return render(

        request,

        "dashboard/reports.html",

        {

            "revenue_by_payment": revenue_by_payment,

            "best_customers": best_customers,

            "low_stock": low_stock,

        },

    )


@owner_or_staff_required
def sales_report_pdf(request):

    orders = (
        Order.objects
        .select_related(
            "user",
            "coupon",
        )
        .order_by("-created_at")
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="sales_report.pdf"'
    )

    doc = SimpleDocTemplate(response)

    data = [[
        "Order",
        "Customer",
        "Date",
        "Payment",
        "Status",
        "Coupon",
        "Discount",
        "Total",
    ]]

    for order in orders:

        data.append([

            f"#{order.id}",

            order.user.username,

            order.created_at.strftime("%d-%m-%Y"),

            order.payment_method.upper(),

            order.status.title(),

            order.coupon.code if order.coupon else "-",

            f"${order.discount}",

            f"${order.final_total}",

        ])

    table = Table(data)

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("GRID",(0,0),(-1,-1),1,colors.grey),

        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

        ("BACKGROUND",(0,1),(-1,-1),colors.whitesmoke),

    ]))

    doc.build([table])

    return response


@owner_or_staff_required
def sales_report_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Sales Report"

    sheet.append([

        "Order ID",

        "Customer",

        "Date",

        "Payment",

        "Status",

        "Coupon",

        "Discount",

        "Total",

    ])

    orders = (
        Order.objects
        .select_related(
            "user",
            "coupon",
        )
        .order_by("-created_at")
    )

    for order in orders:

        sheet.append([

            order.id,

            order.user.username,

            order.created_at.strftime("%d-%m-%Y"),

            order.payment_method,

            order.status,

            order.coupon.code if order.coupon else "",

            float(order.discount),

            float(order.final_total),

        ])

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        'attachment; filename="sales_report.xlsx"'

    )

    workbook.save(response)

    return response


def low_stock_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="low_stock_report.pdf"'

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    data = [

        [

            "Product",

            "Category",

            "Stock",

        ]

    ]

    products = Product.objects.filter(

        stock__lte=5

    ).select_related(

        "category"

    )

    for product in products:

        data.append([

            product.name,

            product.category.name,

            str(product.stock),

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("GRID",(0,0),(-1,-1),1,colors.grey),

            ("BACKGROUND",(0,0),(-1,0),colors.black),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("BOTTOMPADDING",(0,0),(-1,0),8),

        ])

    )

    elements = [

        Paragraph(

            "Low Stock Report",

            styles["Heading1"]

        ),

        table,

    ]

    doc.build(elements)

    return response


def low_stock_report_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Low Stock"

    ws.append([

        "Product",

        "Category",

        "Stock",

    ])

    products = Product.objects.filter(

        stock__lte=5

    ).select_related(

        "category"

    )

    for product in products:

        ws.append([

            product.name,

            product.category.name,

            product.stock,

        ])

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response[
        "Content-Disposition"

    ] = 'attachment; filename="low_stock.xlsx"'

    wb.save(response)

    return response


@owner_required
def category_list(request):

    categories = Category.objects.only(
        "id",
        "name",
    ).order_by("name")

    return render(
        request,
        "dashboard/categories/list.html",
        {
            "categories": categories,
        },
    )


@owner_required
def category_create(request):

    if request.method == "POST":

        form = CategoryForm(request.POST, request.FILES)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Category created successfully.",
            )

            return redirect("dashboard:category_list")

    else:

        form = CategoryForm()

    return render(
        request,
        "dashboard/categories/create.html",
        {
            "form": form,
        },
    )


@owner_required
def category_update(request, pk):

    category = get_object_or_404(
        Category,
        pk=pk,
    )

    if request.method == "POST":

        form = CategoryForm(
            request.POST,
            request.FILES,
            instance=category,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Category updated successfully.",
            )

            return redirect("dashboard:category_list")

    else:

        form = CategoryForm(instance=category)

    return render(
        request,
        "dashboard/categories/update.html",
        {
            "form": form,
            "category": category,
        },
    )


@owner_required
def category_delete(request, pk):

    category = get_object_or_404(
        Category,
        pk=pk,
    )

    category.delete()

    messages.success(
        request,
        "Category deleted successfully.",
    )

    return redirect("dashboard:category_list")


# ==========================================
# Brand List
# ==========================================

def brand_list(request):

    brands = Brand.objects.order_by("name")

    return render(
        request,
        "dashboard/brands/list.html",
        {
            "brands": brands,
        },
    )


# ==========================================
# Brand Create
# ==========================================

def brand_create(request):

    if request.method == "POST":

        form = BrandForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Brand Added Successfully.",
            )

            return redirect("dashboard:brand_list")

    else:

        form = BrandForm()

    return render(
        request,
        "dashboard/brands/create.html",
        {
            "form": form,
        },
    )


# ==========================================
# Brand Update
# ==========================================

def brand_update(request, pk):

    brand = get_object_or_404(
        Brand,
        pk=pk,
    )

    if request.method == "POST":

        form = BrandForm(
            request.POST,
            request.FILES,
            instance=brand,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Brand Updated Successfully.",
            )

            return redirect("dashboard:brand_list")

    else:

        form = BrandForm(
            instance=brand,
        )

    return render(
        request,
        "dashboard/brands/update.html",
        {
            "form": form,
        },
    )


# ==========================================
# Brand Delete
# ==========================================

def brand_delete(request, pk):

    brand = get_object_or_404(Brand, pk=pk)

    if request.method == "POST":

        brand.delete()

        messages.success(
            request,
            "Brand Deleted Successfully.",
        )

        return redirect("dashboard:brand_list")

    return render(
        request,
        "dashboard/brands/delete.html",
        {
            "brand": brand,
        },
    )


# ==========================================
# Product List
# ==========================================
@owner_required



def product_list(request):

    query = request.GET.get("q")

    products = (
        Product.objects
        .select_related(
            "category",
            "brand",
        )
        .order_by("-created_at")
    )

    if query:

        products = products.filter(
            name__icontains=query
        )

    paginator = Paginator(
        products,
        10,
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {

        "page_obj": page_obj,

        "products": page_obj,

        "query": query,

        "total_products": Product.objects.count(),

        "available_products": Product.objects.filter(
            available=True
        ).count(),

        "featured_products": Product.objects.filter(
            featured=True
        ).count(),

        "out_of_stock": Product.objects.filter(
            stock=0
        ).count(),

    }

    return render(
        request,
        "dashboard/products/list.html",
        context,
    )

# ==========================================
# Product Create
# ==========================================

@owner_required
def product_create(request):


    if request.method == "POST":

        form = ProductForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():

            product = form.save()

            for image in request.FILES.getlist("images"):

                ProductImage.objects.create(
                    product=product,
                    image=image,
                )

            messages.success(
                    request,
                "Product created successfully.",
            )

            return redirect("dashboard:product_list")

    else:

        form = ProductForm()

        

    return render(
        request,
        "dashboard/products/create.html",
        {
            "form": form,
        },
    )

# ==========================================
# Product Update
# ==========================================

@owner_required
def product_update(request, pk):

    product = get_object_or_404(
        Product,
        pk=pk,
    )

    if request.method == "POST":

        form = ProductForm(
            request.POST,
            request.FILES,
            instance=product,
        )

        if form.is_valid():

            product = form.save()

            for image in request.FILES.getlist("images"):

                ProductImage.objects.create(
                    product=product,
                    image=image,
                )
                

            messages.success(
                request,
                "Product updated successfully.",
            )

            return redirect("dashboard:product_list")

    else:

        form = ProductForm(
            instance=product,
        )

        

    return render(
        request,
        "dashboard/products/update.html",
        {
            "form": form,
            "product": product,
            "gallery": product.gallery.all(),
        },
    )


# ==========================================
# Product Delete
# ==========================================
@owner_required
def product_delete(request, pk):

    product = get_object_or_404(
        Product,
        pk=pk,
    )

    if request.method == "POST":

        product.delete()

        messages.success(
            request,
            "Product deleted successfully.",
        )

        return redirect("dashboard:product_list")

    return render(
        request,
        "dashboard/products/delete.html",
        {

            "product": product,

        },
    )



# ==========================================
# Product Gallery
# ==========================================

def product_gallery(request, pk):

    product = get_object_or_404(
        Product,
        pk=pk,
    )

    images = product.gallery.only(
        "id",
        "image",
    )

    form = ProductGalleryForm()

    context = {

        "product": product,

        "images": images,

        "form": form,

    }

    return render(

        request,

        "dashboard/product_images/gallery.html",

        context,

    )



# ==========================================
# Upload Product Image
# ==========================================

def product_image_create(request, pk):

    product = get_object_or_404(

        Product,

        pk=pk,

    )

    if request.method == "POST":

        form = ProductGalleryForm(

            request.POST,

            request.FILES,

        )

        if form.is_valid():

            image = form.save(

                commit=False,

            )

            image.product = product

            image.save()

            messages.success(

                request,

                "Image Uploaded Successfully.",

            )

    return redirect(

        "dashboard:product_gallery",

        pk=pk,

    )


# ==========================================
# Delete Product Image
# ==========================================

@owner_required
def product_image_delete(request, pk):

    image = get_object_or_404(
        ProductImage,
        pk=pk,
    )

    product_id = image.product.id

    image.delete()

    messages.success(
        request,
        "Image deleted successfully.",
    )

    return redirect(
        "dashboard:product_update",
        pk=product_id,
    )


# ==========================================
# Coupon List
# ==========================================

def coupon_list(request):

    coupons = Coupon.objects.only(
        "id",
        "code",
        "discount",
        "active",
        "valid_from",
        "valid_to",
    ).order_by("-id")

    return render(

        request,

        "dashboard/coupon/list.html",

        {

            "coupons": coupons,

        },

    )


# ==========================================
# Coupon Add
# ==========================================

def coupon_add(request):

    if request.method == "POST":

        form = CouponForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect("dashboard:coupon_list")

    else:

        form = CouponForm()

    return render(

        request,

        "dashboard/coupon/form.html",

        {

            "form": form,

            "title": "Add Coupon",

        },

    )


# ==========================================
# Coupon Edit
# ==========================================

def coupon_edit(request, pk):

    coupon = get_object_or_404(

        Coupon,

        pk=pk,

    )

    if request.method == "POST":

        form = CouponForm(

            request.POST,

            instance=coupon,

        )

        if form.is_valid():

            form.save()

            return redirect(

                "dashboard:coupon_list"

            )

    else:

        form = CouponForm(

            instance=coupon,

        )

    return render(

        request,

        "dashboard/coupon/form.html",

        {

            "form": form,

            "title": "Edit Coupon",

        },

    )


# ==========================================
# Coupon Delete
# ==========================================

def coupon_delete(request, pk):

    coupon = get_object_or_404(

        Coupon,

        pk=pk,

    )

    coupon.delete()

    return redirect(

        "dashboard:coupon_list"

    )

@owner_or_staff_required
def dashboard_order_detail(request, order_id):

    order = get_object_or_404(
        Order.objects.select_related(
            "user",
            "coupon",
        ).prefetch_related(
            "items__product",
            "timeline__user",
        ),
        id=order_id,
    )

    if request.method == "POST":

        # ====================================================
        # Add Timeline Only
        # ====================================================

        if "note" in request.POST and request.POST.get("note").strip():

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=request.POST.get("note"),

            )

            messages.success(
                request,
                "Timeline note added successfully."
            )

            return redirect(
                "dashboard:dashboard_order_detail",
                order.id,
            )

        # ====================================================
        # Previous Values
        # ====================================================

        old_status = order.status
        old_payment = order.payment_status

        old_tracking = order.tracking_number
        old_courier = order.courier_name
        old_delivery = order.estimated_delivery

        # ====================================================
        # Payment
        # ====================================================

        order.payment_status = request.POST.get(
            "payment_status",
            order.payment_status,
        )

        order.paid = order.payment_status in [
            "paid",
            "partial",
        ]

        # ====================================================
        # Order Status
        # ====================================================

        order.status = request.POST.get(
            "status",
            order.status,
        )

        # ====================================================
        # Courier
        # ====================================================

        order.courier_name = request.POST.get(
            "courier_name",
            "",
        )

        order.delivery_charge = (
            request.POST.get("delivery_charge")
            or 0
        )

        estimated = request.POST.get(
            "estimated_delivery"
        )

        if estimated:

            order.estimated_delivery = estimated

        # ====================================================
        # Save
        # ====================================================

        order.save()
        order.refresh_from_db()

        # ====================================================
        # Return Request Update
        # ====================================================

        if hasattr(order, "return_request"):

            request_obj = order.return_request

            old_return_status = request_obj.status

            return_status = request.POST.get(
                "return_status",
                request_obj.status,
            )

            resolution = request.POST.get(
                "resolution",
                request_obj.resolution,
            )

            request_obj.status = return_status
            request_obj.resolution = resolution


            request_obj.save()

            # ==========================================
# Return Comment
# ==========================================

            new_admin_note = request.POST.get(
                "return_admin_note",
                "",
            ).strip()

            customer_visible = (
                request.POST.get("customer_visible") == "on"
            )

            if new_admin_note:

                ReturnComment.objects.create(

                    return_request=request_obj,

                    user=request.user,

                    message=new_admin_note,

                    customer_visible=customer_visible,

                )

                OrderTimeline.objects.create(

                    order=order,

                    user=request.user,

                    created_by=request.user.username,

                    note="💬 Admin replied to return request.",

                )

            if old_return_status != return_status:

                OrderTimeline.objects.create(
                    order=order,
                    user=request.user,
                    created_by=request.user.username,
                    note=(
                        f"↩ Return request updated "
                        f"({old_return_status} → {return_status})"
                    ),
                )

    # ==========================================
    # Approved
    # ==========================================

            if return_status == "approved":

                first_item = order.items.select_related(
                    "product"
                ).first()

        # ---------------- Refund ----------------

                if resolution == "refund":

                    Refund.objects.get_or_create(
                        return_request=request_obj,
                        defaults={
                            "order": order,
                            "amount": order.final_total,
                            "status": "pending",
                        },
                    )

                    OrderTimeline.objects.create(
                        order=order,
                        user=request.user,
                        created_by=request.user.username,
                        note="💰 Refund request sent to Refund Center.",
                    )

        # ---------------- Exchange ----------------

                elif resolution == "exchange":

                    if first_item:

                        Exchange.objects.get_or_create(
                            return_request=request_obj,
                            defaults={
                                "order": order,
                                "old_product": first_item.product,
                                "status": "pending",
                            },
                        )

                        OrderTimeline.objects.create(
                            order=order,
                            user=request.user,
                            created_by=request.user.username,
                            note="🔄 Exchange request sent to Exchange Center.",
                    )

        # ---------------- Replacement ----------------

                elif resolution == "replacement":

                    if first_item:

                        Replacement.objects.get_or_create(
                            return_request=request_obj,
                            defaults={
                                "order": order,
                                "product": first_item.product,
                                "quantity": first_item.quantity,
                                "status": "pending",
                            },
                        )

                        OrderTimeline.objects.create(
                            order=order,
                            user=request.user,
                            created_by=request.user.username,
                            note="📦 Replacement request sent to Replacement Center.",
                        )

        # ====================================================
        # Tracking Timeline
        # ====================================================

        if old_tracking != order.tracking_number and order.tracking_number:

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=f"📍 Tracking Number Added : {order.tracking_number}",

            )


        if old_courier != order.courier_name and order.courier_name:

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=f"🚚 Courier Assigned : {order.courier_name}",

            )


        if old_delivery != order.estimated_delivery and order.estimated_delivery:

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=f"📅 Estimated Delivery : {order.estimated_delivery}",

            )

        # ====================================================
        # Admin Note -> Timeline
        # ====================================================

        admin_note = request.POST.get("admin_note")

        if admin_note:

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=admin_note,

            )

        # ====================================================
        # Timeline Auto Log
        # ====================================================

        if old_status != order.status:

            STATUS_MESSAGES = {
                "pending": "📦 Order placed.",
                "processing": "⚙️ Order is being processed.",
                "shipped": f"🚚 Order shipped via {order.courier_name}.",
                "delivered": "✅ Order delivered successfully.",
                "cancelled": "❌ Order cancelled.",
            }

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=STATUS_MESSAGES.get(order.status, order.status),

            )

        if old_payment != order.payment_status:

            PAYMENT_MESSAGES = {
                "pending": "💳 Payment Pending.",
                "partial": "💰 Partial Payment Received.",
                "paid": "✅ Payment Completed.",
                "failed": "❌ Payment Failed.",
                "refunded": "↩️ Payment Refunded.",
            }

            OrderTimeline.objects.create(

                order=order,

                user=request.user,

                created_by=request.user.username,

                note=PAYMENT_MESSAGES.get(order.payment_status),

            )

        # ====================================================
        # Email
        # ====================================================

        from payments.utils import (
            send_shipping_email,
            send_delivered_email,
            send_cancelled_email,
        )

        # Send Shipping Email
        if (
            order.status == "shipped"
            and (
                old_status != order.status
                or old_courier != order.courier_name
                or old_tracking != order.tracking_number
                or old_delivery != order.estimated_delivery
            )
        ):
            send_shipping_email(request, order)

        # Delivered Email
        if (
            old_status != order.status
                and order.status == "delivered"
        ):
            send_delivered_email(request, order)

        # Cancelled Email
        if (
            old_status != order.status
                and order.status == "cancelled"
            ):
            send_cancelled_email(request, order)

        messages.success(

            request,

            "Order updated successfully."

        )

        return redirect(

            "dashboard:dashboard_order_detail",

            order.id,

        )

    return render(

        request,

        "dashboard/order_detail.html",

        {

            "order": order,

        },

    )


@owner_or_staff_required
def dashboard_orders(request):

    orders = Order.objects.select_related(
        "user"
    ).order_by(
        "-created_at"
    )

    # ==========================
    # Search
    # ==========================

    q = request.GET.get("q")

    if q:

        orders = orders.filter(

            Q(id__icontains=q)

            |

            Q(full_name__icontains=q)

            |

            Q(email__icontains=q)

            |

            Q(phone__icontains=q)

        )

    # ==========================
    # Status
    # ==========================

    status = request.GET.get("status")

    if status:

        orders = orders.filter(
            status=status
        )

    # ==========================
    # Payment Method
    # ==========================

    payment = request.GET.get("payment")

    if payment:

        orders = orders.filter(
            payment_method=payment
        )

    # ==========================
    # Paid
    # ==========================

    paid = request.GET.get("paid")

    if paid == "yes":

        orders = orders.filter(
            paid=True
        )

    elif paid == "no":

        orders = orders.filter(
            paid=False
        )

    # ==========================
    # Pagination
    # ==========================

    paginator = Paginator(
        orders,
        20,
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(

        request,

        "dashboard/orders.html",

        {

            "orders": page_obj,

            "page_obj": page_obj,

        },

    )


@owner_or_staff_required
def export_orders_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Orders"

    ws.append([
        "Order ID",
        "Customer",
        "Email",
        "Phone",
        "Total",
        "Payment",
        "Payment Status",
        "Order Status",
        "Courier",
        "Tracking",
        "Date",
    ])

    orders = Order.objects.select_related(
        "user"
    ).order_by(
        "-created_at"
    )

    for order in orders:

        ws.append([

            order.id,

            order.full_name,

            order.email,

            order.phone,

            float(order.final_total),

            order.payment_method,

            order.payment_status,

            order.status,

            order.courier_name,

            order.tracking_number,

            order.created_at.strftime("%d-%m-%Y %H:%M"),

        ])

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (
        'attachment; filename="orders.xlsx"'
    )

    wb.save(response)

    return response

@owner_or_staff_required
def low_stock_report(request):

    products = Product.objects.filter(stock__lte=5)

    return render(
        request,
        "dashboard/low_stock_report.html",
        {
            "products": products,
        },
    )


@owner_or_staff_required
def dashboard_customers(request):

    q = request.GET.get("q")

    customers = (
        User.objects
        .filter(role="customer")
        .annotate(
            total_orders=Count("orders", distinct=True),
            total_spent=Sum("orders__final_total"),
        )
        .order_by("-date_joined")
    )

    if q:

        customers = customers.filter(
            Q(username__icontains=q)
            |
            Q(email__icontains=q)
            |
            Q(phone__icontains=q)
        )

    paginator = Paginator(customers, 20)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    # Stats

    now = timezone.now()

    online_threshold = now - timedelta(minutes=5)

    context = {

        "customers": page_obj,
        "page_obj": page_obj,

        "total_customers": User.objects.filter(
            role="customer"
        ).count(),

        "online_customers": User.objects.filter(
            role="customer",
            last_activity__gte=online_threshold,
            is_blocked=False,
        ).count(),

        "blocked_customers": User.objects.filter(
            role="customer",
            is_blocked=True,
        ).count(),

        "today_joined": User.objects.filter(
            role="customer",
            date_joined__date=now.date(),
        ).count(),
    }

    return render(
        request,
        "dashboard/customers.html",
        context,
    )


@owner_or_staff_required
def dashboard_customer_detail(request, user_id):

    customer = get_object_or_404(
        User,
        id=user_id,
        role="customer",
    )

    orders = customer.orders.select_related(
        "coupon",
    ).order_by(
        "-created_at",
    )

    total_spent = (
        orders.filter(
            paid=True,
            status="delivered",
        ).aggregate(
            total=Sum("final_total"),
        )["total"] or 0
    )

    # Online status

    is_online = False

    if customer.last_activity:

        is_online = (
            timezone.now() - customer.last_activity
        ).total_seconds() <= 300  # 5 minutes

    context = {

        "customer": customer,

        "orders": orders,

        "total_spent": total_spent,

        "is_online": is_online,

    }

    return render(
        request,
        "dashboard/customer_detail.html",
        context,
    )


@owner_or_staff_required
def customer_block(request, user_id):

    customer = get_object_or_404(
        User,
        id=user_id,
        role="customer",
    )

    customer.is_blocked = True
    customer.save(update_fields=["is_blocked"])

    messages.success(
        request,
        f"{customer.username} has been blocked.",
    )

    return redirect(
        "dashboard:dashboard_customer_detail",
        user_id=customer.id,
    )


@owner_or_staff_required
def customer_unblock(request, user_id):

    customer = get_object_or_404(
        User,
        id=user_id,
        role="customer",
    )

    customer.is_blocked = False
    customer.save(update_fields=["is_blocked"])

    messages.success(
        request,
        f"{customer.username} has been unblocked.",
    )

    return redirect(
        "dashboard:dashboard_customer_detail",
        user_id=customer.id,
    )


@owner_or_staff_required
def customer_activity(request):

    activities = (
        CustomerActivity.objects
        .select_related(
            "user",
            "product",
        )
        .order_by("-created_at")
    )

    q = request.GET.get("q")
    action = request.GET.get("action")

    if q:
        activities = activities.filter(
            Q(user__username__icontains=q)
            |
            Q(user__email__icontains=q)
            |
            Q(product__name__icontains=q)
        )

    if action:
        activities = activities.filter(
            action=action,
        )

    paginator = Paginator(
        activities,
        30,
    )

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    context = {

        "page_obj": page_obj,

        "cart_count": CustomerActivity.objects.filter(
            action="cart_add",
        ).count(),

        "wishlist_count": CustomerActivity.objects.filter(
            action="wishlist_add",
        ).count(),

        "today_count": CustomerActivity.objects.filter(
            created_at__date=timezone.now().date(),
        ).count(),

    }

    return render(
        request,
        "dashboard/customer_activity.html",
        context,
    )

@owner_or_staff_required
def online_customers(request):

    online_threshold = timezone.now() - timedelta(minutes=5)

    customers = (
        User.objects.filter(
            role="customer",
            is_blocked=False,
        )
        .only(
            "id",
            "username",
            "email",
            "phone",
            "profile_image",
            "last_activity",
            "last_seen_page",
            "browser",
            "last_ip",
            "is_blocked",
        )
        .order_by("-last_activity")
    )

    context = {

        "customers": customers,

        "online_threshold": online_threshold,

        "online_count": customers.filter(
            last_activity__gte=online_threshold,
        ).count(),

    }

    return render(
        request,
        "dashboard/online_customers.html",
        context,
    )


@owner_or_staff_required
def abandoned_carts(request):

    carts = (
        AbandonedCart.objects
        .select_related(
            "user",
            "product",
        )
        .filter(
            recovered=False,
        )
        .order_by(
            "-updated_at",
        )
    )

    context = {

        "carts": carts,

    }

    return render(
        request,
        "dashboard/abandoned_carts.html",
        context,
    )

@owner_or_staff_required
def send_abandoned_cart_email(request, cart_id):

    cart = get_object_or_404(
        AbandonedCart.objects.select_related(
            "user",
            "product",
        ),
        id=cart_id,
    )

    customer = cart.user
    product = cart.product

    message = f"""
Hi {customer.first_name or customer.username},

You left the following product in your cart.

Product:
{product.name}

Quantity:
{cart.quantity}

Complete your order before it goes out of stock.

Thank you.
"""

    email = EmailMultiAlternatives(
        subject=f"Complete your order - {product.name}",
        body=message,
        to=[customer.email],
    )

    email.send()

    cart.reminder_sent = True

    # যদি last_reminder field add করে থাকো
    # cart.last_reminder = timezone.now()

    cart.save(
        update_fields=[
            "reminder_sent",
            # "last_reminder",
        ]
    )

    messages.success(
        request,
        "Reminder email sent successfully.",
    )

    return redirect(
        "dashboard:abandoned_carts",
    )

@owner_or_staff_required
def wishlist_recovery(request):

    items = Wishlist.objects.select_related(
        "user",
        "product",
    ).order_by(
        "-created_at",
    )

    q = request.GET.get("q")

    if q:

        items = items.filter(
            Q(user__username__icontains=q)
            |
            Q(user__email__icontains=q)
            |
            Q(product__name__icontains=q)
        )

    paginator = Paginator(items, 30)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {

        "page_obj": page_obj,

        "total_items": Wishlist.objects.count(),

        "pending_items": Wishlist.objects.filter(
            promotion_sent=False,
        ).count(),

        "sent_items": Wishlist.objects.filter(
            promotion_sent=True,
        ).count(),
    }

    return render(
        request,
        "dashboard/wishlist_recovery.html",
        context,
    )

@owner_or_staff_required
def send_wishlist_promotion(request, wishlist_id):

    wishlist = get_object_or_404(
        Wishlist.objects.select_related(
            "user",
            "product",
        ),
        id=wishlist_id,
    )

    product = wishlist.product
    user = wishlist.user

    message = f"""
Hi {user.first_name or user.username},

Good news!

Your wishlist product is still available.

Product:
{product.name}

Price:
${product.current_price}

Visit now and complete your purchase.

Thank you.
"""

    send_mail(
        subject=f"{product.name} is waiting for you ❤️",
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=False,
    )

    wishlist.promotion_sent = True
    wishlist.last_promotion = timezone.now()

    wishlist.save(
        update_fields=[
            "promotion_sent",
            "last_promotion",
        ]
    )

    messages.success(
        request,
        "Promotion email sent successfully.",
    )

    return redirect(
        "dashboard:wishlist_recovery",
    )

@owner_or_staff_required
def currency_exchange(request):

    exchange_rate, created = ExchangeRate.objects.get_or_create(
        currency="USD",
        defaults={
            "rate": 122.50,
        },
    )

    if request.method == "POST":

        exchange_rate.rate = request.POST.get("rate")
        exchange_rate.save()

        messages.success(
            request,
            "USD exchange rate updated successfully.",
        )

        return redirect("dashboard:home")

    return render(
        request,
        "dashboard/currency_exchange.html",
        {
            "exchange_rate": exchange_rate,
        },
    )

@owner_or_staff_required
def feature_toggle_list(request):

    defaults = [

        {
            "key": "flash_sale",
            "category": "marketing",
            "icon": "bi bi-lightning-charge-fill",
            "description": "Enable flash sale products.",
        },

        {
            "key": "free_delivery",
            "category": "marketing",
            "icon": "bi bi-truck",
            "description": "Show free delivery offers.",
        },

        {
            "key": "new_arrival",
            "category": "products",
            "icon": "bi bi-stars",
            "description": "Display new arrival section.",
        },

        {
            "key": "trending",
            "category": "products",
            "icon": "bi bi-fire",
            "description": "Display trending products.",
        },

        {
            "key": "guest_checkout",
            "category": "checkout",
            "icon": "bi bi-person",
            "description": "Allow guest checkout.",
        },

        {
            "key": "wishlist",
            "category": "website",
            "icon": "bi bi-heart",
            "description": "Enable wishlist feature.",
        },

        {
            "key": "reviews",
            "category": "website",
            "icon": "bi bi-chat-left-text",
            "description": "Allow product reviews.",
        },

        {
            "key": "coupon",
            "category": "checkout",
            "icon": "bi bi-ticket",
            "description": "Enable coupon system.",
        },

        {
            "key": "cod",
            "category": "payment",
            "icon": "bi bi-cash-stack",
            "description": "Cash On Delivery payment.",
        },

        {
            "key": "emi",
            "category": "payment",
            "icon": "bi bi-credit-card",
            "description": "Enable EMI payment.",
        },

        {
            "key": "maintenance",
            "category": "website",
            "icon": "bi bi-tools",
            "description": "Put website into maintenance mode.",
        },

        {
            "key": "coming_soon",
            "category": "website",
            "icon": "bi bi-hourglass-split",
            "description": "Show coming soon page.",
        },

        {
            "key": "announcement",
            "category": "marketing",
            "icon": "bi bi-megaphone",
            "description": "Display announcement bar.",
        },

    ]

    for item in defaults:

        FeatureToggle.objects.update_or_create(

            key=item["key"],

            defaults={

                "category": item["category"],

                "icon": item["icon"],

                "description": item["description"],

            },

        )

    if request.method == "POST":

        feature = get_object_or_404(
            FeatureToggle,
            id=request.POST.get("id"),
        )

        feature.enabled = not feature.enabled

        feature.updated_by = request.user

        feature.save()

        clear_feature_cache()

        messages.success(
            request,
            f"{feature.get_key_display()} updated successfully.",
        )

        return redirect("dashboard:feature_toggle_list")

    category = request.GET.get("category")

    features = FeatureToggle.objects.all()

    if category:
        features = features.filter(category=category)

    return render(

        request,

        "dashboard/feature_toggle_list.html",

        {

            "features": features,

            "active_category": category,

            "categories": FeatureToggle.CATEGORY_CHOICES,

        },

    )


@owner_required
def website_features(request):

    settings = WebsiteSettings.load()

    if request.method == "POST":

        settings.shop_name = request.POST.get("shop_name")
        settings.shop_tagline = request.POST.get("shop_tagline")

        settings.announcement = request.POST.get("announcement")

        settings.phone = request.POST.get("phone")
        settings.email = request.POST.get("email")
        settings.address = request.POST.get("address")

        settings.facebook = request.POST.get("facebook")
        settings.instagram = request.POST.get("instagram")
        settings.youtube = request.POST.get("youtube")
        settings.linkedin = request.POST.get("linkedin")
        settings.twitter = request.POST.get("twitter")

        settings.footer_text = request.POST.get("footer_text")
        settings.copyright = request.POST.get("copyright")

        if request.FILES.get("logo"):
            settings.logo = request.FILES["logo"]

        if request.FILES.get("favicon"):
            settings.favicon = request.FILES["favicon"]

        settings.save()

        messages.success(request, "Website settings updated successfully.")
        return redirect("dashboard:website_features")

    return render(
        request,
        "dashboard/website_features.html",
        {
            "settings": settings,
        },
    )

@owner_required
def announcement_list(request):

    announcements = Announcement.objects.all()

    return render(
        request,
        "dashboard/announcement/list.html",
        {
            "announcements": announcements,
        },
    )


@owner_required
def announcement_add(request):

    if request.method == "POST":

        form = AnnouncementForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Announcement added successfully.",
            )

            return redirect("dashboard:announcement_list")

    else:

        form = AnnouncementForm()

    return render(
        request,
        "dashboard/announcement/form.html",
        {
            "form": form,
            "title": "Add Announcement",
        },
    )


@owner_required
def announcement_edit(request, pk):

    announcement = get_object_or_404(
        Announcement,
        pk=pk,
    )

    if request.method == "POST":

        form = AnnouncementForm(
            request.POST,
            instance=announcement,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Announcement updated successfully.",
            )

            return redirect("dashboard:announcement_list")

    else:

        form = AnnouncementForm(
            instance=announcement,
        )

    return render(
        request,
        "dashboard/announcement/form.html",
        {
            "form": form,
            "title": "Edit Announcement",
        },
    )

@owner_required
def announcement_delete(request, pk):

    announcement = get_object_or_404(
        Announcement,
        pk=pk,
    )

    if request.method == "POST":

        announcement.delete()

        messages.success(
            request,
            "Announcement deleted successfully.",
        )

        return redirect("dashboard:announcement_list")

    return render(
        request,
        "dashboard/announcement/delete.html",
        {
            "announcement": announcement,
        },
    )

@owner_or_staff_required
def dashboard_returns(request):

    status = request.GET.get("status")

    returns = (
        ReturnRequest.objects
        .select_related("order", "user")
        .order_by("-created_at")
    )

    if status:
        returns = returns.filter(status=status)

    paginator = Paginator(returns, 10)   # প্রতি পেজে 10টি

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    context = {
        "returns": page_obj,

        "page_obj": page_obj,

        "pending_count": ReturnRequest.objects.filter(status="pending").count(),
        "approved_count": ReturnRequest.objects.filter(status="approved").count(),
        "rejected_count": ReturnRequest.objects.filter(status="rejected").count(),
        "refunded_count": ReturnRequest.objects.filter(status="refunded").count(),
        "all_count": ReturnRequest.objects.count(),
    }

    return render(
        request,
        "dashboard/returns.html",
        context,
    )


@owner_or_staff_required
def dashboard_refunds(request):

    status = request.GET.get("status")

    refunds = (
        Refund.objects
        .select_related(
            "order",
            "return_request",
            "completed_by",
        )
        .order_by("-created_at")
    )

    if status:

        refunds = refunds.filter(
            status=status,
        )

    paginator = Paginator(
        refunds,
        20,
    )

    page = request.GET.get("page")

    page_obj = paginator.get_page(page)

    context = {

        "refunds": page_obj,

        "page_obj": page_obj,

        "pending_count": Refund.objects.filter(
            status="pending",
        ).count(),

        "processing_count": Refund.objects.filter(
            status="processing",
        ).count(),

        "completed_count": Refund.objects.filter(
            status="completed",
        ).count(),

        "cancelled_count": Refund.objects.filter(
            status="cancelled",
        ).count(),

        "total_amount": (
            Refund.objects.filter(
                status="completed",
            ).aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        ),

        "current_status": status,

    }

    return render(

        request,

        "dashboard/refunds.html",

        context,

    )



@owner_or_staff_required
def dashboard_refund_detail(request, pk):

    refund = get_object_or_404(
        Refund.objects.select_related(
            "order",
            "return_request",
        ),
        pk=pk,
    )

    return render(
        request,
        "dashboard/refund_detail.html",
        {
            "refund": refund,
        },
    )

@require_POST
@owner_or_staff_required
def complete_refund(request, pk):

    refund = get_object_or_404(
        Refund.objects.select_related(
            "order",
            "return_request",
        ),
        pk=pk,
    )

    # Already completed
    if refund.status == "completed":

        messages.warning(
            request,
            "Refund already completed.",
        )

        return redirect(
            "dashboard:dashboard_refund_detail",
            refund.id,
        )

    # Cancelled refund cannot be completed
    if refund.status == "cancelled":

        messages.error(
            request,
            "Cancelled refund cannot be completed.",
        )

        return redirect(
            "dashboard:dashboard_refund_detail",
            refund.id,
        )

    # ==========================================
    # Refund
    # ==========================================

    refund.status = "completed"
    refund.completed_at = timezone.now()
    refund.completed_by = request.user

    refund.admin_note = request.POST.get(
        "admin_note",
        refund.admin_note,
    )

    refund.refund_method = request.POST.get(
        "refund_method",
        refund.refund_method,
    )

    refund.transaction_id = request.POST.get(
        "transaction_id",
        refund.transaction_id,
    )

    refund.save()

    # ==========================================
    # Order
    # ==========================================

    order = refund.order

    order.payment_status = "refunded"
    order.status = "cancelled"

    order.save(
        update_fields=[
            "payment_status",
            "status",
        ]
    )

    # ==========================================
    # Return Request
    # ==========================================

    refund.return_request.status = "refunded"
    refund.return_request.admin_note = refund.admin_note

    refund.return_request.save(
        update_fields=[
            "status",
            "admin_note",
            "updated_at",
        ]
    )

    # ==========================================
    # Timeline
    # ==========================================

    OrderTimeline.objects.create(
        order=order,
        user=request.user,
        created_by=request.user.username,
        note=(
            "💰 Refund completed successfully.\n\n"
            f"Refund Method : {refund.get_refund_method_display()}\n"
            f"Transaction ID : {refund.transaction_id or 'N/A'}"
        ),
    )

    messages.success(
        request,
        "Refund completed successfully.",
    )

    return redirect(
        "dashboard:dashboard_refund_detail",
        refund.id,
    )

@require_POST
@owner_or_staff_required
def cancel_refund(request, pk):

    refund = get_object_or_404(
        Refund.objects.select_related(
            "order",
            "return_request",
        ),
        pk=pk,
    )

    # Already completed
    if refund.status == "completed":

        messages.error(
            request,
            "Completed refund cannot be cancelled.",
        )

        return redirect(
            "dashboard:dashboard_refund_detail",
            refund.id,
        )

    refund.status = "cancelled"

    refund.admin_note = request.POST.get(
        "admin_note",
        refund.admin_note,
    )

    refund.save(
        update_fields=[
            "status",
            "admin_note",
        ]
    )

    refund.return_request.status = "approved"
    refund.return_request.admin_note = refund.admin_note

    refund.return_request.save(
        update_fields=[
            "status",
            "admin_note",
            "updated_at",
        ]
    )

    OrderTimeline.objects.create(
        order=refund.order,
        user=request.user,
        created_by=request.user.username,
        note="❌ Refund request cancelled by admin.",
    )

    messages.success(
        request,
        "Refund cancelled successfully.",
    )

    return redirect(
        "dashboard:dashboard_refund_detail",
        refund.id,
    )

@owner_or_staff_required
def dashboard_exchanges(request):

    status = request.GET.get("status")

    exchanges = (
        Exchange.objects
        .select_related(
            "order",
            "old_product",
            "new_product",
            "return_request",
            "completed_by",
        )
        .order_by("-created_at")
    )

    if status:

        exchanges = exchanges.filter(
            status=status,
        )

    paginator = Paginator(
        exchanges,
        20,
    )

    page = request.GET.get("page")

    page_obj = paginator.get_page(page)

    context = {

        "exchanges": page_obj,

        "page_obj": page_obj,

        "pending_count": Exchange.objects.filter(
            status="pending",
        ).count(),

        "approved_count": Exchange.objects.filter(
            status="approved",
        ).count(),

        "completed_count": Exchange.objects.filter(
            status="completed",
        ).count(),

        "cancelled_count": Exchange.objects.filter(
            status="cancelled",
        ).count(),

        "current_status": status,

    }

    return render(

        request,

        "dashboard/exchanges.html",

        context,

    )



@owner_or_staff_required
def dashboard_exchange_detail(request, pk):

    exchange = get_object_or_404(
        Exchange.objects.select_related(
            "order",
            "old_product",
            "new_product",
            "return_request",
        ),
        pk=pk,
    )

    if request.method == "POST":

        product_id = request.POST.get(
            "new_product",
        )

        if product_id:

            exchange.new_product = get_object_or_404(
                Product,
                pk=product_id,
            )

        exchange.status = request.POST.get(
            "status",
            exchange.status,
        )

        exchange.admin_note = request.POST.get(
            "admin_note",
            exchange.admin_note,
        )

        exchange.save()

        OrderTimeline.objects.create(

            order=exchange.order,

            user=request.user,

            created_by=request.user.username,

            note=(
                f"🔄 Exchange updated.\n"
                f"Status : {exchange.get_status_display()}"
            ),

        )

        messages.success(

            request,

            "Exchange updated successfully.",

        )

        return redirect(

            "dashboard:dashboard_exchange_detail",

            exchange.id,

        )

    products = Product.objects.filter(
    available=True,
    ).order_by(
        "name",
    )

    return render(

        request,

        "dashboard/exchange_detail.html",

        {
            "exchange": exchange,
            "products": products,
        },

    )


@require_POST
@owner_or_staff_required
def complete_exchange(request, pk):

    exchange = get_object_or_404(
        Exchange.objects.select_related(
            "order",
            "old_product",
            "new_product",
            "return_request",
        ).prefetch_related(
            "order__items__product",
        ),
        pk=pk,
    )

    if exchange.status == "completed":

        messages.warning(
            request,
            "Exchange already completed.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    if exchange.status == "cancelled":

        messages.error(
            request,
            "Cancelled exchange cannot be completed.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    if not exchange.new_product:

        messages.error(
            request,
            "Please select a new product before completing the exchange.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    # ==========================================
    # Order Item & Quantity
    # ==========================================

    order_item = exchange.order.items.select_related(
        "product"
    ).first()

    if not order_item:

        messages.error(
            request,
            "Order item not found.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    quantity = order_item.quantity

    # ==========================================
    # Stock Check
    # ==========================================

    if exchange.new_product.stock < quantity:

        messages.error(
            request,
            "Insufficient stock for the selected product.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    # ==========================================
    # Complete Exchange
    # ==========================================

    exchange.status = "completed"
    exchange.completed_by = request.user
    exchange.completed_at = timezone.now()
    exchange.admin_note = request.POST.get(
        "admin_note",
        exchange.admin_note,
    )

    exchange.save()

    # ==========================================
    # Restore Old Product Stock
    # ==========================================

    exchange.old_product.stock += quantity

    exchange.old_product.save(
        update_fields=[
            "stock",
        ]
    )

    # ==========================================
    # Reduce New Product Stock
    # ==========================================

    exchange.new_product.stock -= quantity

    exchange.new_product.save(
        update_fields=[
            "stock",
        ]
    )

    # ==========================================
    # Update Return Request
    # ==========================================

    exchange.return_request.status = "completed"
    exchange.return_request.admin_note = exchange.admin_note

    exchange.return_request.save(
        update_fields=[
            "status",
            "admin_note",
        ]
    )

    # ==========================================
    # Timeline
    # ==========================================

    OrderTimeline.objects.create(

        order=exchange.order,

        user=request.user,

        created_by=request.user.username,

        note=(
            f"🔄 Exchange completed.\n"
            f"Old Product : {exchange.old_product.name}\n"
            f"New Product : {exchange.new_product.name}\n"
            f"Quantity : {quantity}"
        ),

    )

    messages.success(

        request,

        "Exchange completed successfully.",

    )

    return redirect(

        "dashboard:dashboard_exchange_detail",

        exchange.id,

    )


@require_POST
@owner_or_staff_required
def cancel_exchange(request, pk):

    exchange = get_object_or_404(
        Exchange.objects.select_related(
            "order",
            "return_request",
        ),
        pk=pk,
    )

    if exchange.status == "completed":

        messages.error(
            request,
            "Completed exchange cannot be cancelled.",
        )

        return redirect(
            "dashboard:dashboard_exchange_detail",
            exchange.id,
        )

    exchange.status = "cancelled"

    exchange.admin_note = request.POST.get(
        "admin_note",
        exchange.admin_note,
    )

    exchange.save(
        update_fields=[
            "status",
            "admin_note",
        ]
    )

    exchange.return_request.status = "approved"

    exchange.return_request.admin_note = exchange.admin_note

    exchange.return_request.save(
        update_fields=[
            "status",
            "admin_note",
        ]
    )

    OrderTimeline.objects.create(

        order=exchange.order,

        user=request.user,

        created_by=request.user.username,

        note="❌ Exchange cancelled by admin.",

    )

    messages.success(

        request,

        "Exchange cancelled successfully.",

    )

    return redirect(

        "dashboard:dashboard_exchange_detail",

        exchange.id,

    )

@owner_or_staff_required
def dashboard_replacements(request):

    status = request.GET.get("status")
    q = request.GET.get("q")

    replacements = (
        Replacement.objects
        .select_related(
            "order",
            "product",
            "return_request",
            "completed_by",
        )
        .order_by("-created_at")
    )

    if q:

        replacements = replacements.filter(

            Q(order__id__icontains=q)

            |

            Q(order__full_name__icontains=q)

            |

            Q(order__email__icontains=q)

            |

            Q(product__name__icontains=q)

        )

    if status:

        replacements = replacements.filter(
            status=status
        )

    paginator = Paginator(
        replacements,
        20,
    )

    page = request.GET.get("page")

    page_obj = paginator.get_page(page)

    context = {

        "replacements": page_obj,

        "page_obj": page_obj,

        "pending_count": Replacement.objects.filter(
            status="pending"
        ).count(),

        "approved_count": Replacement.objects.filter(
            status="approved"
        ).count(),

        "completed_count": Replacement.objects.filter(
            status="completed"
        ).count(),

        "cancelled_count": Replacement.objects.filter(
            status="cancelled"
        ).count(),

    }

    return render(
        request,
        "dashboard/replacements.html",
        context,
    )

    
@owner_or_staff_required
def dashboard_replacement_detail(request, pk):

    replacement = get_object_or_404(

        Replacement.objects.select_related(

            "order",

            "product",

            "return_request",

            "completed_by",

        ),

        pk=pk,

    )

    if request.method == "POST":

        replacement.status = request.POST.get(

            "status",

            replacement.status,

        )

        replacement.admin_note = request.POST.get(

            "admin_note",

            replacement.admin_note,

        )

        replacement.save()

        messages.success(

            request,

            "Replacement updated successfully.",

        )

        return redirect(

            "dashboard:dashboard_replacement_detail",

            replacement.id,

        )

    return render(

        request,

        "dashboard/replacement_detail.html",

        {

            "replacement": replacement,

        },

    )

@require_POST
@owner_or_staff_required
def complete_replacement(request, pk):

    replacement = get_object_or_404(
        Replacement.objects.select_related(
            "order",
            "product",
            "return_request",
        ),
        pk=pk,
    )

    if replacement.status == "completed":

        messages.warning(
            request,
            "Replacement already completed.",
        )

        return redirect(
            "dashboard:dashboard_replacement_detail",
            replacement.id,
        )

    # ==========================
    # Update Replacement
    # ==========================

    replacement.status = "completed"

    replacement.completed_by = request.user

    replacement.completed_at = timezone.now()

    replacement.admin_note = request.POST.get(
        "admin_note",
        replacement.admin_note,
    )

    replacement.save()

    # ==========================
    # Update Return Request
    # ==========================

    replacement.return_request.status = "completed"

    replacement.return_request.save(
        update_fields=[
            "status",
        ]
    )

    # ==========================
    # Timeline
    # ==========================

    OrderTimeline.objects.create(

        order=replacement.order,

        user=request.user,

        created_by=request.user.username,

        note=(
            f"📦 Replacement completed.\n"
            f"Product : {replacement.product.name}\n"
            f"Quantity : {replacement.quantity}"
        ),

    )

    messages.success(

        request,

        "Replacement completed successfully.",

    )

    return redirect(

        "dashboard:dashboard_replacement_detail",

        replacement.id,

    )


@require_POST
@owner_or_staff_required
def cancel_replacement(request, pk):

    replacement = get_object_or_404(
        Replacement.objects.select_related(
            "order",
            "return_request",
        ),
        pk=pk,
    )

    if replacement.status == "completed":

        messages.error(

            request,

            "Completed replacement cannot be cancelled.",

        )

        return redirect(

            "dashboard:dashboard_replacement_detail",

            replacement.id,

        )

    replacement.status = "cancelled"

    replacement.admin_note = request.POST.get(

        "admin_note",

        replacement.admin_note,

    )

    replacement.save()

    OrderTimeline.objects.create(

        order=replacement.order,

        user=request.user,

        created_by=request.user.username,

        note="❌ Replacement cancelled.",

    )

    messages.success(

        request,

        "Replacement cancelled successfully.",

    )

    return redirect(

        "dashboard:dashboard_replacement_detail",

        replacement.id,

    )


@owner_or_staff_required
def return_review(request, pk):

    return_request = get_object_or_404(
        ReturnRequest.objects.select_related(
            "order",
            "user",
        ).prefetch_related(
            "order__items__product",
            "order__timeline",
        ),
        pk=pk,
    )

    order = return_request.order

    if request.method == "POST":

        resolution = request.POST.get("resolution")

        admin_note = request.POST.get(
            "admin_note",
            "",
        ).strip()

        if not resolution:

            messages.error(
                request,
                "Please select a resolution.",
            )

            return redirect(
                "dashboard:return_review",
                pk=pk,
            )

        return_request.admin_note = admin_note

        return_request.resolution = resolution

        # =====================================
        # Reject
        # =====================================

        if resolution == "reject":

            return_request.status = "rejected"

            return_request.save()

            OrderTimeline.objects.create(
                order=order,
                user=request.user,
                created_by=request.user.username,
                note="❌ Return request rejected.",
            )

            messages.success(
                request,
                "Return request rejected.",
            )

            return redirect(
                "dashboard:dashboard_returns",
            )

        # =====================================
        # Approved
        # =====================================

        return_request.status = "approved"

        return_request.save()

        OrderTimeline.objects.create(
            order=order,
            user=request.user,
            created_by=request.user.username,
            note=f"✅ Return approved ({resolution.title()}).",
        )

        order_item = order.items.first()

        if not order_item:

            messages.error(
                request,
                "Order item not found.",
            )

            return redirect(
                "dashboard:dashboard_returns",
            )

        # =====================================
        # Refund
        # =====================================

        if resolution == "refund":

            Refund.objects.get_or_create(
                order=order,
                return_request=return_request,
                defaults={
                    "amount": order.final_total,
                    "status": "pending",
                },
            )

            OrderTimeline.objects.create(
                order=order,
                user=request.user,
                created_by=request.user.username,
                note="💰 Sent to Refund Center.",
            )

            messages.success(
                request,
                "Refund request sent to Refund Center.",
            )

        # =====================================
        # Exchange
        # =====================================

        elif resolution == "exchange":

            Exchange.objects.get_or_create(
                order=order,
                return_request=return_request,
                defaults={
                    "old_product": order_item.product,
                    "status": "pending",
                },
            )

            OrderTimeline.objects.create(
                order=order,
                user=request.user,
                created_by=request.user.username,
                note="🔄 Sent to Exchange Center.",
            )

            messages.success(
                request,
                "Exchange request sent to Exchange Center.",
            )

        # =====================================
        # Replacement
        # =====================================

        elif resolution == "replacement":

            Replacement.objects.get_or_create(
                order=order,
                return_request=return_request,
                defaults={
                    "product": order_item.product,
                    "quantity": order_item.quantity,
                    "status": "pending",
                },
            )

            OrderTimeline.objects.create(
                order=order,
                user=request.user,
                created_by=request.user.username,
                note="📦 Sent to Replacement Center.",
            )

            messages.success(
                request,
                "Replacement request sent to Replacement Center.",
            )

        return redirect(
            "dashboard:dashboard_returns",
        )

    return render(
        request,
        "dashboard/return_review.html",
        {
            "return_request": return_request,
            "order": order,
        },
    )



@owner_required
def product_ranking(request):

    products = Product.objects.order_by(
        "display_order",
        "-created_at",
    )

    if request.method == "POST":

        product_id = request.POST.get("product_id")
        new_rank = request.POST.get("display_order")

        try:
            product = Product.objects.get(id=product_id)
            new_rank = int(new_rank)

        except (Product.DoesNotExist, ValueError):

            messages.error(
                request,
                "Invalid request.",
            )

            return redirect(
                "dashboard:product_ranking",
            )

        # ==========================================
        # Validation
        # ==========================================

        max_rank = Product.objects.count()

        if new_rank < 1 or new_rank > max_rank:

            messages.error(
                request,
                f"Rank must be between 1 and {max_rank}.",
            )

            return redirect(
                "dashboard:product_ranking",
            )

        # ==========================================
        # Already Same Rank
        # ==========================================

        if product.display_order == new_rank:

            messages.info(
                request,
                "This product is already in that position.",
            )

            return redirect(
                "dashboard:product_ranking",
            )

        # ==========================================
        # Swap Ranking
        # ==========================================

        with transaction.atomic():

            old_rank = product.display_order

            try:

                target = Product.objects.get(
                    display_order=new_rank
                )

                target.display_order = old_rank

                target.save(
                    update_fields=[
                        "display_order",
                    ]
                )

            except Product.DoesNotExist:

                pass

            product.display_order = new_rank

            product.save(
                update_fields=[
                    "display_order",
                ]
            )

        messages.success(
            request,
            "Product ranking updated successfully.",
        )

        return redirect(
            "dashboard:product_ranking",
        )

    # ==========================================
    # Statistics
    # ==========================================

    total_products = Product.objects.count()

    highest = Product.objects.order_by(
        "display_order",
    ).first()

    lowest = Product.objects.order_by(
        "-display_order",
    ).first()

    return render(
        request,
        "dashboard/product_ranking.html",
        {
            "products": products,
            "total_products": total_products,
            "highest": highest,
            "lowest": lowest,
        },
    )




@owner_or_staff_required
def email_campaign_list(request):

    campaigns = (
        EmailCampaign.objects
        .select_related("created_by")
        .annotate(
            success_count=Count(
                "logs",
                filter=Q(logs__status="success"),
            ),
            failed_count=Count(
                "logs",
                filter=Q(logs__status="failed"),
            ),
            recipient_count=Count("logs"),
        )
        .order_by("-created_at")
    )

    context = {
        "campaigns": campaigns,

        "total_campaigns": campaigns.count(),

        "draft_count": EmailCampaign.objects.filter(
            status="draft"
        ).count(),

        "completed_count": EmailCampaign.objects.filter(
            status="completed"
        ).count(),

        "total_sent": EmailCampaignLog.objects.filter(
            status="success"
        ).count(),
    }

    return render(
        request,
        "dashboard/email_campaign_list.html",
        context,
    )


@owner_or_staff_required
def email_campaign_create(request):

    if request.method == "POST":

        form = EmailCampaignForm(request.POST)

        if form.is_valid():

            campaign = form.save(commit=False)

            campaign.created_by = request.user

            campaign.save()

            messages.success(
                request,
                "Campaign created successfully.",
            )

            return redirect(
                "dashboard:email_campaign_list",
            )

    else:

        form = EmailCampaignForm()

    return render(
        request,
        "dashboard/email_campaign_form.html",
        {
            "form": form,
        },
    )


@owner_or_staff_required
def email_campaign_send(request, campaign_id):

    campaign = get_object_or_404(
        EmailCampaign,
        id=campaign_id,
    )

    campaign.status = "sending"
    campaign.save(update_fields=["status"])

    # -----------------------------
    # Target Users
    # -----------------------------

    if campaign.target == "all":

        customers = User.objects.filter(
            role="customer",
            is_active=True,
            is_blocked=False,
        )

    elif campaign.target == "wishlist":

        customers = User.objects.filter(
            wishlist__isnull=False,
            role="customer",
            is_blocked=False,
        ).distinct()

    elif campaign.target == "cart":

        customers = User.objects.filter(
            abandonedcart__recovered=False,
            role="customer",
            is_blocked=False,
        ).distinct()

    elif campaign.target == "no_order":

        customers = User.objects.filter(
            role="customer",
            orders__isnull=True,
            is_blocked=False,
        ).distinct()

    else:

        customers = User.objects.none()

    # -----------------------------
    # Send Email
    # -----------------------------

    for customer in customers:

        try:

            email = EmailMultiAlternatives(

                subject=campaign.subject,

                body=campaign.message,

                to=[customer.email],

            )

            email.send()

            EmailCampaignLog.objects.create(

                campaign=campaign,

                customer=customer,

                status="success",

            )

        except Exception as e:

            EmailCampaignLog.objects.create(

                campaign=campaign,

                customer=customer,

                status="failed",

                error=str(e),

            )

    campaign.status = "completed"

    campaign.sent_at = timezone.now()

    campaign.save(
        update_fields=[
            "status",
            "sent_at",
        ]
    )

    messages.success(
        request,
        "Campaign sent successfully.",
    )

    return redirect(
        "dashboard:email_campaign_list",
    )


@owner_or_staff_required
def email_campaign_edit(request, campaign_id):

    campaign = get_object_or_404(
        EmailCampaign,
        id=campaign_id,
    )

    if request.method == "POST":

        form = EmailCampaignForm(
            request.POST,
            instance=campaign,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Campaign updated successfully.",
            )

            return redirect(
                "dashboard:email_campaign_list",
            )

    else:

        form = EmailCampaignForm(
            instance=campaign,
        )

    return render(
        request,
        "dashboard/email_campaign_form.html",
        {
            "form": form,
            "campaign": campaign,
        },
    )


@owner_or_staff_required
def email_campaign_delete(request, campaign_id):

    campaign = get_object_or_404(
        EmailCampaign,
        id=campaign_id,
    )

    campaign.delete()

    messages.success(
        request,
        "Campaign deleted successfully.",
    )

    return redirect(
        "dashboard:email_campaign_list",
    )